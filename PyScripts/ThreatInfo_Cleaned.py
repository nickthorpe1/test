import requests
import json
import pandas as pd
from openpyxl import load_workbook

# Set your Mandiant API credentials here
API_ID = ''
API_SECRET = ''

AUTH_URL = 'https://api.intelligence.mandiant.com/token'
BASE_URL = 'https://api.intelligence.mandiant.com/v4'
EXCEL_FILE = 'threat_actor_info_Cleaned.xlsx'


# Function to get the authentication token
def get_auth_token(api_id, api_secret):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
    }
    data = {
        'grant_type': 'client_credentials'
    }

    try:
        response = requests.post(AUTH_URL, headers=headers, data=data, auth=(api_id, api_secret))
        response.raise_for_status()  # Raise an exception for HTTP errors

        token = response.json().get('access_token')
        if not token:
            raise ValueError("Failed to retrieve access token from response")

        return token

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
        print(f"Response content: {response.content}")
    except Exception as err:
        print(f"An error occurred: {err}")


# Function to get threat actor information
def get_threat_actor_info(token, actor_name):
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json'
    }

    # Construct the endpoint URL
    endpoint = f'{BASE_URL}/actor/{actor_name}'

    try:
        # Make the API request
        response = requests.get(endpoint, headers=headers)
        response.raise_for_status()  # Raise an exception for HTTP errors

        # Parse the JSON response
        data = response.json()

        # Print the results
        print(json.dumps(data, indent=4))

        return data

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
        print(f"Response content: {response.content}")
    except Exception as err:
        print(f"An error occurred: {err}")


# Function to format nested dictionaries into strings
def format_nested_data(data):
    if isinstance(data, list):
        formatted_data = []
        for item in data:
            formatted_str = ""
            for key, value in item.items():
                formatted_str += f"{key.capitalize()}: {value}\n"
            formatted_data.append(formatted_str.strip())
        return "\n\n".join(formatted_data)
    return data


# Function to save data to Excel
def save_to_excel(data, actor_name):
    # Flatten the JSON data and format nested dictionaries
    flat_data = pd.json_normalize(data)
    for col in flat_data.columns:
        if isinstance(flat_data[col].iloc[0], list):
            flat_data[col] = flat_data[col].apply(format_nested_data)

    # Load existing workbook or create a new one if it doesn't exist
    try:
        book = load_workbook(EXCEL_FILE)
        if actor_name in book.sheetnames:
            existing_df = pd.read_excel(EXCEL_FILE, sheet_name=actor_name)
            last_updated_existing = pd.to_datetime(existing_df['last_updated']).max()

            # Filter the new data based on the last_updated field
            new_data = flat_data[pd.to_datetime(flat_data['last_updated']) > last_updated_existing]

            # If there is new data, append it
            if not new_data.empty:
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    new_data.to_excel(writer, sheet_name=actor_name, index=False, header=False,
                                      startrow=len(existing_df) + 1)
        else:
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a') as writer:
                flat_data.to_excel(writer, sheet_name=actor_name, index=False, header=True)
    except FileNotFoundError:
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            flat_data.to_excel(writer, sheet_name=actor_name, index=False, header=True)
    except Exception as e:
        print(f"An error occurred while saving to Excel: {e}")


# Example usage
if __name__ == "__main__":
    actor_names = "UNC2165, TEMP.Armageddon, UNC5812, UNC4697, APT44, UNC2926, UNC4515, UNC4814, UNC5435, UNC5175, UNC4895, UNC4221, UNC5716, Turla Team, UNC3707, UNC3628, UNC4057, UNC5464, UNC5227, UNC638, UNC5417, UNC5544, UNC5101, UNC5125, APT29, UNC5252, UNC5271, UNC3507, UNC4940, UNC5658, UNC2589, UNC4322, UNC2386, UNC3735, APT28, UNC4027, UNC2524, UNC2628, UNC2680, TEMP.Isotope, UNC753, ZeroWing Team, UNC5114, Koala Team"  # Replace with comma-separated names of the threat actors you are interested in
    actor_list = [actor.strip() for actor in actor_names.split(",")]

    # Get the authentication token
    token = get_auth_token(API_ID, API_SECRET)
    if token:
        for actor_name in actor_list:
            # Get the threat actor information
            data = get_threat_actor_info(token, actor_name)
            if data:
                # Save the threat actor information to Excel
                save_to_excel(data, actor_name)
